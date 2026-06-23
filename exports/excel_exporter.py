"""
Excel Exporter — produces branded two-sheet workbooks:
  Sheet 1: Filtered data with ISH header and formatted table
  Sheet 2: Filter metadata (report name, filters applied, timestamp)
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Brand tokens ─────────────────────────────────────────────────────────────
NAVY       = "0D2137"
GOLD       = "C9A84C"
WHITE      = "FFFFFF"
LIGHT_BLUE = "EAF0F7"
MID_BLUE   = "D0DCE8"
GREY_TEXT  = "5A6A7A"
BORDER_CLR = "B0BEC5"

FONT_BODY  = "Arial"
FONT_TITLE = "Arial"


def _thin_border(color=BORDER_CLR):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill():
    return PatternFill("solid", fgColor=NAVY)


def _alt_fill():
    return PatternFill("solid", fgColor=LIGHT_BLUE)


def export_to_excel(
    df: pd.DataFrame,
    report_label: str,
    active_filters: dict,
) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Data ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Report Data"

    _write_data_sheet(ws, df, report_label, active_filters)

    # ── Sheet 2: Filter Metadata ─────────────────────────────────────────
    wm = wb.create_sheet("Filter Metadata")
    _write_metadata_sheet(wm, report_label, active_filters)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_data_sheet(ws, df: pd.DataFrame, report_label: str, active_filters: dict):
    # ── Title block ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 4))
    title_cell = ws.cell(row=1, column=1, value="ISH Information Systems")
    title_cell.font = Font(name=FONT_TITLE, bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(df.columns), 4))
    sub_cell = ws.cell(row=2, column=1, value=f"Report: {report_label}")
    sub_cell.font = Font(name=FONT_TITLE, bold=False, size=10, color=WHITE)
    sub_cell.fill = PatternFill("solid", fgColor=NAVY)
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    ts_cell = ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    ts_cell.font = Font(name=FONT_BODY, size=9, italic=True, color=GREY_TEXT)
    ws.row_dimensions[3].height = 14
    ws.cell(row=4, column=1)  # spacer
    ws.row_dimensions[4].height = 6

    # ── Column headers ────────────────────────────────────────────────────
    header_row = 5
    display_df = _prep_display_df(df)

    for col_idx, col_name in enumerate(display_df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(name=FONT_BODY, bold=True, size=10, color=WHITE)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border(NAVY)
    ws.row_dimensions[header_row].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, row in enumerate(display_df.itertuples(index=False), start=header_row + 1):
        fill = _alt_fill() if (row_idx - header_row) % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_clean_val(value))
            cell.font = Font(name=FONT_BODY, size=9)
            cell.alignment = Alignment(vertical="center", indent=1)
            cell.border = _thin_border()
            if fill:
                cell.fill = fill
            # Right-align numerics
            if isinstance(value, (int, float)) and not pd.isna(value):
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row_idx].height = 15

    # ── Auto-fit columns ─────────────────────────────────────────────────
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        col_vals = [str(col_name)] + [
            str(_clean_val(v)) for v in display_df.iloc[:, col_idx - 1]
        ]
        max_len = max((len(v) for v in col_vals), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # ── Freeze panes ─────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_metadata_sheet(ws, report_label: str, active_filters: dict):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    def _hdr(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=FONT_BODY, bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell.alignment = Alignment(indent=1)

    def _row(ws, r, label, value):
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=str(value))
        a.font = Font(name=FONT_BODY, bold=True, size=9, color=GREY_TEXT)
        b.font = Font(name=FONT_BODY, size=9)
        a.fill = b.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for cell in (a, b):
            cell.border = _thin_border()
            cell.alignment = Alignment(indent=1, vertical="center")
        ws.row_dimensions[r].height = 15

    _hdr(1, "ISH Information Systems — Export Metadata")
    ws.row_dimensions[1].height = 20

    _row(ws, 2, "Report", report_label)
    _row(ws, 3, "Generated At", datetime.now().strftime("%d %b %Y, %I:%M %p"))
    _row(ws, 4, "Active Filter Count", len(active_filters) if active_filters else "None")

    _hdr(6, "Filters Applied")
    ws.row_dimensions[6].height = 18

    if active_filters:
        for i, (k, v) in enumerate(active_filters.items(), start=7):
            val_str = ", ".join(v) if isinstance(v, list) else str(v)
            _row(ws, i, k, val_str)
    else:
        cell = ws.cell(row=7, column=1, value="No filters applied — full dataset exported.")
        cell.font = Font(name=FONT_BODY, size=9, italic=True, color=GREY_TEXT)
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)


def _prep_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Drop internal helper columns (prefixed _) before export."""
    drop_cols = [c for c in df.columns if c.startswith("_") or c in ("SNAP Test 1", "SNAP Test 2", "SNAP Test 3")]
    return df.drop(columns=[c for c in drop_cols if c in df.columns])


def _clean_val(val):
    if pd.isna(val) if not isinstance(val, (list, dict)) else False:
        return None
    if isinstance(val, bool):
        return "Yes" if val else "No"
    return val
